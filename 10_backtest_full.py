"""
10_backtest_full.py — Walk-forward backtest 2015-2025
For each season: train on all prior seasons, predict current season.
No data leakage — all features in match_features.csv are expanding-window.
Output: data/backtest_full.xlsx with 5 sheets.
"""

import pandas as pd
import numpy as np
import pickle
from pathlib import Path
from collections import defaultdict

import xgboost as xgb
import optuna
from sklearn.model_selection import KFold, cross_val_predict
from sklearn.metrics import log_loss, accuracy_score
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

optuna.logging.set_verbosity(optuna.logging.WARNING)
Path("data").mkdir(exist_ok=True)

TEAM_NORM = {
    "Delhi Daredevils": "Delhi Capitals",
    "Deccan Chargers": "Sunrisers Hyderabad",
    "Rising Pune Supergiants": "Rising Pune Supergiant",
    "Punjab Kings": "Kings XI Punjab",
    "Kings XI Punjab": "Kings XI Punjab",
    "Royal Challengers Bangalore": "Royal Challengers Bengaluru",
    "Royal Challengers Bengaluru": "Royal Challengers Bengaluru",
}

# ── Load features ─────────────────────────────────────────────────────────
print("Loading match_features.csv...")
df = pd.read_csv("data/match_features.csv", parse_dates=["date"])
df = df[df["team1_won"].notna()].copy()
df["season"] = df["season"].astype(str)
df["team1"] = df["team1"].map(lambda x: TEAM_NORM.get(x, x))
df["team2"] = df["team2"].map(lambda x: TEAM_NORM.get(x, x))

ALL_FEATURES = [
    "elo_diff",
    "team1_form", "team2_form", "form_diff",
    "form_3_diff", "form_10_diff", "form_weighted_diff",
    "h2h_win_rate_team1",
    "team1_won_toss", "toss_chose_bat",
    "venue_toss_win_rate", "venue_bat_first_win_rate",
    "venue_avg_first_innings", "venue_chase_win_rate",
    "team1_venue_win_rate", "team2_venue_win_rate",
    "match_num_in_season", "is_playoff",
    "bat_diff", "bowl_diff",
    "team1_bat_strength", "team2_bat_strength",
    "team1_bowl_strength", "team2_bowl_strength",
    "team1_bats_second", "toss_venue_aligned",
    "team1_chase_wr", "team2_chase_wr", "chase_wr_diff",
    "team1_chase_advantage", "team2_chase_advantage", "chase_advantage_diff",
    "early_season", "early_chase_boost",
    "venue_chase_batting_second",
]
ALL_FEATURES = [f for f in ALL_FEATURES if f in df.columns]
print(f"Using {len(ALL_FEATURES)} features")

# ── Tune hyperparameters ONCE on pre-2015 data ────────────────────────────
print("\nTuning XGBoost on pre-2015 data (40 trials)...")
pre2015 = df[df["season"] < "2015"].copy()
train_median_global = pre2015[ALL_FEATURES].median()

X_tune = pre2015[ALL_FEATURES].fillna(train_median_global)
y_tune = pre2015["team1_won"].astype(int)
kf = KFold(n_splits=5, shuffle=False)

def objective(trial):
    params = {
        "n_estimators":      trial.suggest_int("n_estimators", 50, 400),
        "max_depth":         trial.suggest_int("max_depth", 2, 5),
        "learning_rate":     trial.suggest_float("learning_rate", 0.01, 0.2, log=True),
        "subsample":         trial.suggest_float("subsample", 0.6, 1.0),
        "colsample_bytree":  trial.suggest_float("colsample_bytree", 0.4, 1.0),
        "min_child_weight":  trial.suggest_int("min_child_weight", 1, 10),
        "reg_alpha":         trial.suggest_float("reg_alpha", 1e-4, 10.0, log=True),
        "reg_lambda":        trial.suggest_float("reg_lambda", 1e-4, 10.0, log=True),
        "gamma":             trial.suggest_float("gamma", 0, 2.0),
    }
    model = xgb.XGBClassifier(**params, eval_metric="logloss",
                               random_state=42, verbosity=0)
    oof = cross_val_predict(model, X_tune, y_tune, cv=kf,
                            method="predict_proba")[:, 1]
    return log_loss(y_tune, oof)

study = optuna.create_study(direction="minimize")
study.optimize(objective, n_trials=40)
BEST_PARAMS = study.best_params
print(f"  Best CV log-loss: {study.best_value:.4f}")
print(f"  Params: {BEST_PARAMS}")

# ── Walk-forward backtest 2015-2025 ───────────────────────────────────────
TEST_SEASONS = [s for s in sorted(df["season"].unique()) if s >= "2015"]
all_results = []

for season in TEST_SEASONS:
    train = df[df["season"] < season].copy()
    test  = df[df["season"] == season].copy()

    if len(train) < 50 or len(test) == 0:
        print(f"  Skipping {season} (insufficient data)")
        continue

    train_median = train[ALL_FEATURES].median()
    X_tr = train[ALL_FEATURES].fillna(train_median)
    y_tr = train["team1_won"].astype(int)
    X_te = test[ALL_FEATURES].fillna(train_median)
    y_te = test["team1_won"].astype(int)

    model = xgb.XGBClassifier(**BEST_PARAMS, eval_metric="logloss",
                               random_state=42, verbosity=0)
    model.fit(X_tr, y_tr)
    probs = model.predict_proba(X_te)[:, 1]
    preds = (probs >= 0.5).astype(int)

    acc = accuracy_score(y_te, preds)
    print(f"  {season}: {acc*100:.1f}%  ({preds.sum()} predicted T1, {(1-preds).sum()} T2)  [{len(test)} matches]")

    test = test.copy()
    test["prob_team1"]       = probs
    test["predicted_winner"] = np.where(preds == 1, test["team1"], test["team2"])
    test["correct"]          = (preds == y_te.values).astype(int)
    all_results.append(test)

results = pd.concat(all_results, ignore_index=True)
print(f"\nTotal matches backtested: {len(results)}")
print(f"Overall accuracy: {results['correct'].mean()*100:.1f}%")

# ── Enrich with analysis columns ──────────────────────────────────────────
results["day_of_week"]       = results["date"].dt.day_name()
results["day_type"]          = results["date"].dt.dayofweek.map(
    lambda d: "Weekend" if d >= 5 else "Weekday"
)

# First half vs second half: based on match_num_in_season vs season median
season_midpoints = (
    results.groupby("season")["match_num_in_season"].median().rename("mid")
)
results = results.merge(season_midpoints, on="season")
results["tournament_half"] = np.where(
    results["match_num_in_season"] <= results["mid"],
    "First Half", "Second Half"
)
results = results.drop(columns=["mid"])

results["actual_winner"] = np.where(
    results["team1_won"] == 1,
    results["team1"], results["team2"]
)
results["result"] = results["correct"].map({1: "Correct", 0: "Wrong"})

# ── Build Excel ───────────────────────────────────────────────────────────
print("\nBuilding Excel file...")

# ── Helper styles ─────────────────────────────────────────────────────────
def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell_border():
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def pct_bar(val):
    """Return a simple bar string for readability."""
    blocks = int(val / 5)
    return "█" * blocks

HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN    = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center")

GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
GREY_FILL   = PatternFill("solid", fgColor="F2F2F2")
BLUE_HDR    = header_fill("1F4E79")
TEAL_HDR    = header_fill("005F73")
PURPLE_HDR  = header_fill("3D2B8F")
ORANGE_HDR  = header_fill("C05200")
DARK_HDR    = header_fill("2C3E50")

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def write_header_row(ws, row_num, headers, fill, font=None):
    fnt = font or HEADER_FONT
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=ci, value=h)
        c.fill = fill
        c.font = fnt
        c.alignment = HEADER_ALIGN
        c.border = cell_border()

def write_data_row(ws, row_num, values, align="center", alt=False):
    bg = GREY_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row_num, column=ci, value=v)
        c.alignment = CELL_ALIGN if align == "center" else LEFT_ALIGN
        c.border = cell_border()
        c.fill = bg

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — All Games
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "All Games"
ws1.freeze_panes = "A3"
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20

# Title row
ws1.merge_cells("A1:M1")
title_cell = ws1["A1"]
title_cell.value = "IPL MODEL BACKTEST — 2015 to 2025 | Walk-Forward (Train on Prior Seasons)"
title_cell.font = Font(bold=True, color="FFFFFF", size=13)
title_cell.fill = BLUE_HDR
title_cell.alignment = HEADER_ALIGN

hdrs1 = [
    "Date", "Season", "Match #", "Team 1", "Team 2", "Venue",
    "Day", "Day Type", "Tournament Half",
    "T1 Win Prob", "Predicted Winner", "Actual Winner", "Result"
]
write_header_row(ws1, 2, hdrs1, BLUE_HDR)

col_widths1 = [13, 8, 9, 26, 26, 38, 12, 11, 15, 12, 26, 26, 10]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

for ri, (_, row) in enumerate(results.iterrows(), 3):
    alt = (ri % 2 == 0)
    bg  = GREY_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    vals = [
        row["date"].strftime("%d %b %Y"),
        row["season"],
        int(row["match_num_in_season"]),
        row["team1"],
        row["team2"],
        row.get("venue", ""),
        row["day_of_week"],
        row["day_type"],
        row["tournament_half"],
        f"{row['prob_team1']*100:.1f}%",
        row["predicted_winner"],
        row["actual_winner"],
        row["result"],
    ]
    for ci, v in enumerate(vals, 1):
        c = ws1.cell(row=ri, column=ci, value=v)
        c.border = cell_border()
        c.fill = bg
        c.alignment = CELL_ALIGN
        # Colour result column
        if ci == 13:
            c.fill = GREEN_FILL if row["correct"] == 1 else RED_FILL
            c.font = Font(bold=True,
                          color="375623" if row["correct"] == 1 else "9C0006")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — Per Year
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Per Year")
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "ACCURACY BY SEASON — Walk-Forward Backtest"
t2.font = Font(bold=True, color="FFFFFF", size=13)
t2.fill = TEAL_HDR
t2.alignment = HEADER_ALIGN
ws2.row_dimensions[1].height = 30

hdrs2 = ["Season", "Matches", "Correct", "Wrong", "Accuracy %",
         "Weekday Acc %", "Weekend Acc %"]
write_header_row(ws2, 2, hdrs2, TEAL_HDR)

col_widths2 = [10, 10, 10, 10, 14, 16, 16]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

yr_grp = results.groupby("season")
ri = 3
totals = {"m": 0, "c": 0}
for season, grp in yr_grp:
    wkday = grp[grp["day_type"] == "Weekday"]
    wkend = grp[grp["day_type"] == "Weekend"]
    m  = len(grp)
    c  = grp["correct"].sum()
    wd = wkday["correct"].mean() * 100 if len(wkday) > 0 else 0
    we = wkend["correct"].mean() * 100 if len(wkend) > 0 else 0
    acc = c / m * 100
    totals["m"] += m
    totals["c"] += c
    alt = (ri % 2 == 0)
    bg  = GREY_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    row_vals = [season, m, int(c), m - int(c), round(acc, 1), round(wd, 1), round(we, 1)]
    for ci, v in enumerate(row_vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.border = cell_border()
        cell.fill = bg
        cell.alignment = CELL_ALIGN
        if ci == 5:
            if acc >= 60:
                cell.fill = GREEN_FILL
                cell.font = Font(bold=True, color="375623")
            elif acc < 50:
                cell.fill = RED_FILL
                cell.font = Font(bold=True, color="9C0006")
            else:
                cell.font = Font(bold=True)
    ri += 1

# Totals row
tot_acc = totals["c"] / totals["m"] * 100
tot_vals = ["TOTAL", totals["m"], totals["c"],
            totals["m"] - totals["c"], round(tot_acc, 1), "", ""]
for ci, v in enumerate(tot_vals, 1):
    cell = ws2.cell(row=ri, column=ci, value=v)
    cell.border = cell_border()
    cell.fill = header_fill("2C3E50")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = CELL_ALIGN

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — Per Team (Home & Away)
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Per Team")
ws3.freeze_panes = "A3"

ws3.merge_cells("A1:K1")
t3 = ws3["A1"]
t3.value = "ACCURACY BY TEAM — Home Games (Team1) vs Away Games (Team2)"
t3.font = Font(bold=True, color="FFFFFF", size=13)
t3.fill = PURPLE_HDR
t3.alignment = HEADER_ALIGN
ws3.row_dimensions[1].height = 30

hdrs3 = ["Team", "Home M", "Home Correct", "Home Acc %",
         "Away M", "Away Correct", "Away Acc %",
         "Total M", "Total Correct", "Total Acc %", "Best Season"]
write_header_row(ws3, 2, hdrs3, PURPLE_HDR)

col_widths3 = [30, 9, 13, 13, 9, 13, 13, 9, 13, 13, 12]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# team1 = home, team2 = away in match_features.csv perspective
team_stats = defaultdict(lambda: {"hm": 0, "hc": 0, "am": 0, "ac": 0})
for _, row in results.iterrows():
    t1, t2 = row["team1"], row["team2"]
    correct = row["correct"]
    team_stats[t1]["hm"] += 1
    team_stats[t1]["hc"] += correct
    team_stats[t2]["am"] += 1
    # For team2, correct means team2 won (predicted winner was team2)
    team_stats[t2]["ac"] += (1 - correct)

# Best season per team
team_season_acc = (
    results.assign(correct_t1=results["correct"],
                   correct_t2=1 - results["correct"])
    .melt(id_vars=["season", "team1", "team2", "correct"],
          value_vars=["team1"])
    .copy()
)

# Simpler: per team per season accuracy
team_seas = []
for _, row in results.iterrows():
    team_seas.append({"team": row["team1"], "season": row["season"],
                      "correct": row["correct"]})
    team_seas.append({"team": row["team2"], "season": row["season"],
                      "correct": 1 - row["correct"]})
ts_df = pd.DataFrame(team_seas)
best_season = (
    ts_df.groupby(["team", "season"])["correct"].mean()
    .reset_index()
    .sort_values("correct", ascending=False)
    .groupby("team").first()["season"]
)

all_teams = sorted(team_stats.keys())
ri = 3
for team in all_teams:
    s  = team_stats[team]
    hm, hc = s["hm"], s["hc"]
    am, ac = s["am"], s["ac"]
    tm, tc = hm + am, hc + ac
    h_acc  = hc / hm * 100 if hm > 0 else 0
    a_acc  = ac / am * 100 if am > 0 else 0
    t_acc  = tc / tm * 100 if tm > 0 else 0
    bseas  = best_season.get(team, "-")
    alt = (ri % 2 == 0)
    bg  = GREY_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    row_vals = [team, hm, int(hc), round(h_acc, 1),
                am, int(ac), round(a_acc, 1),
                tm, int(tc), round(t_acc, 1), bseas]
    for ci, v in enumerate(row_vals, 1):
        cell = ws3.cell(row=ri, column=ci, value=v)
        cell.border = cell_border()
        cell.fill = bg
        cell.alignment = LEFT_ALIGN if ci == 1 else CELL_ALIGN
        if ci in (4, 7, 10):  # accuracy columns
            if isinstance(v, (int, float)):
                if v >= 60:
                    cell.fill = GREEN_FILL
                    cell.font = Font(bold=True, color="375623")
                elif v < 50:
                    cell.fill = RED_FILL
                    cell.font = Font(bold=True, color="9C0006")
                else:
                    cell.font = Font(bold=True)
    ri += 1

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — Per Venue
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Per Venue")
ws4.freeze_panes = "A3"

ws4.merge_cells("A1:H1")
t4 = ws4["A1"]
t4.value = "ACCURACY BY VENUE — 2015 to 2025"
t4.font = Font(bold=True, color="FFFFFF", size=13)
t4.fill = ORANGE_HDR
t4.alignment = HEADER_ALIGN
ws4.row_dimensions[1].height = 30

hdrs4 = ["Venue", "Matches", "Correct", "Wrong", "Accuracy %",
         "Weekday Acc %", "Weekend Acc %", "First Half Acc %"]
write_header_row(ws4, 2, hdrs4, ORANGE_HDR)

col_widths4 = [42, 10, 10, 10, 14, 15, 15, 16]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

venue_grp = results.groupby("venue").filter(lambda x: len(x) >= 5)
venue_grp = results[results["venue"].isin(
    results.groupby("venue").size()[results.groupby("venue").size() >= 5].index
)]

ri = 3
for venue, grp in venue_grp.groupby("venue"):
    wkday = grp[grp["day_type"] == "Weekday"]
    wkend = grp[grp["day_type"] == "Weekend"]
    fhalf = grp[grp["tournament_half"] == "First Half"]
    m  = len(grp)
    c  = grp["correct"].sum()
    acc = c / m * 100
    wd  = wkday["correct"].mean() * 100 if len(wkday) > 0 else 0
    we  = wkend["correct"].mean() * 100 if len(wkend) > 0 else 0
    fh  = fhalf["correct"].mean() * 100 if len(fhalf) > 0 else 0
    alt = (ri % 2 == 0)
    bg  = GREY_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    row_vals = [venue, m, int(c), m - int(c),
                round(acc, 1), round(wd, 1), round(we, 1), round(fh, 1)]
    for ci, v in enumerate(row_vals, 1):
        cell = ws4.cell(row=ri, column=ci, value=v)
        cell.border = cell_border()
        cell.fill = bg
        cell.alignment = LEFT_ALIGN if ci == 1 else CELL_ALIGN
        if ci == 5:
            if acc >= 60:
                cell.fill = GREEN_FILL
                cell.font = Font(bold=True, color="375623")
            elif acc < 50:
                cell.fill = RED_FILL
                cell.font = Font(bold=True, color="9C0006")
            else:
                cell.font = Font(bold=True)
    ri += 1

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5 — Summary Dashboard
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary", 0)  # insert at front

def summary_block(ws, start_row, start_col, title, rows_data, hdr_fill):
    """Write a labelled block of key-value pairs."""
    # Title
    tc = ws.cell(row=start_row, column=start_col, value=title)
    tc.font = Font(bold=True, color="FFFFFF", size=11)
    tc.fill = hdr_fill
    tc.alignment = HEADER_ALIGN
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 1
    )
    for ri_off, (label, value) in enumerate(rows_data, 1):
        lc = ws.cell(row=start_row + ri_off, column=start_col, value=label)
        vc = ws.cell(row=start_row + ri_off, column=start_col + 1, value=value)
        lc.font = Font(bold=True)
        lc.fill = GREY_FILL
        lc.border = cell_border()
        lc.alignment = LEFT_ALIGN
        vc.border = cell_border()
        vc.alignment = CELL_ALIGN
        vc.font = Font(bold=True, color="1F4E79")

# Banner
ws5.merge_cells("A1:H1")
bn = ws5["A1"]
bn.value = "IPL PREDICTION MODEL — BACKTEST SUMMARY DASHBOARD (2015-2025)"
bn.font = Font(bold=True, color="FFFFFF", size=14)
bn.fill = BLUE_HDR
bn.alignment = HEADER_ALIGN
ws5.row_dimensions[1].height = 36

overall_acc   = results["correct"].mean() * 100
wkday_acc     = results[results["day_type"] == "Weekday"]["correct"].mean() * 100
wkend_acc     = results[results["day_type"] == "Weekend"]["correct"].mean() * 100
fhalf_acc     = results[results["tournament_half"] == "First Half"]["correct"].mean() * 100
shalf_acc     = results[results["tournament_half"] == "Second Half"]["correct"].mean() * 100
best_yr       = results.groupby("season")["correct"].mean().idxmax()
best_yr_acc   = results.groupby("season")["correct"].mean().max() * 100
worst_yr      = results.groupby("season")["correct"].mean().idxmin()
worst_yr_acc  = results.groupby("season")["correct"].mean().min() * 100
best_venue    = (
    results.groupby("venue").filter(lambda x: len(x) >= 10)
    .groupby("venue")["correct"].mean()
)
best_v_name   = best_venue.idxmax()
best_v_acc    = best_venue.max() * 100
worst_v_name  = best_venue.idxmin()
worst_v_acc   = best_venue.min() * 100

overall_data = [
    ("Total Matches", len(results)),
    ("Correct Predictions", int(results["correct"].sum())),
    ("Wrong Predictions",   len(results) - int(results["correct"].sum())),
    ("Overall Accuracy",    f"{overall_acc:.1f}%"),
    ("Seasons Tested",      f"2015 – 2025"),
    ("Method",              "Walk-Forward (No Leakage)"),
]
summary_block(ws5, 3, 1, "OVERALL PERFORMANCE", overall_data, BLUE_HDR)

split_data = [
    ("Weekday Accuracy",         f"{wkday_acc:.1f}%"),
    ("Weekend Accuracy",         f"{wkend_acc:.1f}%"),
    ("First Half Accuracy",      f"{fhalf_acc:.1f}%"),
    ("Second Half Accuracy",     f"{shalf_acc:.1f}%"),
    ("Best Season",              f"{best_yr}  ({best_yr_acc:.1f}%)"),
    ("Worst Season",             f"{worst_yr}  ({worst_yr_acc:.1f}%)"),
]
summary_block(ws5, 3, 4, "SPLITS", split_data, TEAL_HDR)

venue_data = [
    ("Best Venue (≥10 matches)",  f"{best_v_name.split(',')[0]}"),
    ("Best Venue Accuracy",       f"{best_v_acc:.1f}%"),
    ("Worst Venue (≥10 matches)", f"{worst_v_name.split(',')[0]}"),
    ("Worst Venue Accuracy",      f"{worst_v_acc:.1f}%"),
    ("Venues Tested",             results["venue"].nunique()),
    ("Teams Tested",              len(all_teams)),
]
summary_block(ws5, 3, 7, "VENUE & TEAM", venue_data, PURPLE_HDR)

# Per-season mini table on summary sheet
ws5.cell(row=12, column=1, value="SEASON BREAKDOWN").font = Font(bold=True, size=11)
yr_hdrs = ["Season", "Matches", "Accuracy", "Weekday", "Weekend", "First Half", "Playoffs"]
for ci, h in enumerate(yr_hdrs, 1):
    c = ws5.cell(row=13, column=ci, value=h)
    c.fill = DARK_HDR
    c.font = HEADER_FONT
    c.alignment = HEADER_ALIGN
    c.border = cell_border()

ri = 14
for season, grp in results.groupby("season"):
    wkday = grp[grp["day_type"] == "Weekday"]
    wkend = grp[grp["day_type"] == "Weekend"]
    fhalf = grp[grp["tournament_half"] == "First Half"]
    play  = grp[grp["is_playoff"] == 1] if "is_playoff" in grp.columns else pd.DataFrame()
    acc   = grp["correct"].mean() * 100
    wd    = wkday["correct"].mean() * 100 if len(wkday) > 0 else 0
    we    = wkend["correct"].mean() * 100 if len(wkend) > 0 else 0
    fh    = fhalf["correct"].mean() * 100 if len(fhalf) > 0 else 0
    pl    = play["correct"].mean() * 100  if len(play)  > 0 else 0
    alt   = (ri % 2 == 0)
    bg    = GREY_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    row_vals = [season, len(grp), f"{acc:.1f}%", f"{wd:.1f}%",
                f"{we:.1f}%", f"{fh:.1f}%",
                f"{pl:.1f}%" if len(play) > 0 else "N/A"]
    for ci, v in enumerate(row_vals, 1):
        cell = ws5.cell(row=ri, column=ci, value=v)
        cell.border = cell_border()
        cell.fill = bg
        cell.alignment = CELL_ALIGN
        if ci == 3:
            a = float(v.replace("%", ""))
            if a >= 60:
                cell.fill = GREEN_FILL
                cell.font = Font(bold=True, color="375623")
            elif a < 50:
                cell.fill = RED_FILL
                cell.font = Font(bold=True, color="9C0006")
            else:
                cell.font = Font(bold=True)
    ri += 1

col_widths5 = [10, 10, 12, 12, 12, 13, 12, 32, 14, 32, 14]
for i, w in enumerate(col_widths5, 1):
    if i <= ws5.max_column:
        ws5.column_dimensions[get_column_letter(i)].width = w

# Set tab colors
ws5.sheet_properties.tabColor = "1F4E79"
ws1.sheet_properties.tabColor = "2E86AB"
ws2.sheet_properties.tabColor = "005F73"
ws3.sheet_properties.tabColor = "3D2B8F"
ws4.sheet_properties.tabColor = "C05200"

out_path = "data/backtest_full.xlsx"
wb.save(out_path)
print(f"\nSaved: {out_path}")
print(f"  Sheet 1 (Summary):   Overall + per-season breakdown")
print(f"  Sheet 2 (All Games): {len(results)} rows, one per match")
print(f"  Sheet 3 (Per Year):  Season accuracy + weekday/weekend split")
print(f"  Sheet 4 (Per Team):  Home/away accuracy per team")
print(f"  Sheet 5 (Per Venue): Accuracy per venue")
