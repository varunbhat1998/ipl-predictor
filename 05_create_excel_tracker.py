"""
05_create_excel_tracker.py
Creates ipl_predictions.xlsx with proper structure.
Run once before the season starts.

Sheet structure (one row per match):
  MATCH INFO | PRE-MATCH PREDICTION | LIVE PREDICTIONS (over 1..20) | ACTUAL RESULT | ACCURACY
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from pathlib import Path
import datetime

OUTPUT_PATH = "ipl_predictions.xlsx"

# ── Colours ────────────────────────────────────────────────────────────────
C_HEADER_DARK   = "1A3A5C"   # dark navy
C_HEADER_MID    = "2E6DA4"   # mid blue  — match info
C_HEADER_GREEN  = "1D6B3A"   # dark green — pre-match prediction
C_HEADER_ORANGE = "B85C00"   # orange    — live over columns
C_HEADER_PURPLE = "4B0082"   # purple    — actual result
C_HEADER_TEAL   = "0D6B6B"   # teal      — accuracy
C_WHITE         = "FFFFFF"
C_LIGHT_BLUE    = "DCE6F1"
C_LIGHT_GREEN   = "E2EFDA"
C_LIGHT_ORANGE  = "FCE4D6"
C_LIGHT_PURPLE  = "EAD1DC"
C_LIGHT_TEAL    = "D9EAD3"
C_ROW_ALT       = "F5F5F5"

thin = Side(style="thin", color="CCCCCC")
medium = Side(style="medium", color="999999")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
section_border = Border(left=medium, right=medium, top=medium, bottom=medium)

def hdr(text, bg, bold=True, font_color=C_WHITE, size=9, wrap=True, center=True):
    return {
        "value": text, "bg": bg, "bold": bold,
        "font_color": font_color, "size": size,
        "wrap": wrap, "center": center
    }

def apply_cell(cell, value, bg=None, bold=False, font_color="000000",
               size=9, wrap=True, center=True, border=None, number_format=None):
    cell.value = value
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", bold=bold, color=font_color, size=size)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap
    )
    cell.border = border or thin_border
    if number_format:
        cell.number_format = number_format

# ── Column definitions ─────────────────────────────────────────────────────
# Section: MATCH INFO (cols 1–8)
MATCH_COLS = [
    ("match_id",       "Match ID",          8),
    ("season",         "Season",            7),
    ("date",           "Date",             11),
    ("team1",          "Team 1",           22),
    ("team2",          "Team 2",           22),
    ("venue",          "Venue",            26),
    ("toss_winner",    "Toss Winner",      22),
    ("toss_decision",  "Toss Decision",    10),
]

# Section: PRE-MATCH PREDICTION (cols 9–16)
PRE_COLS = [
    ("pre_team1_pct",  "Team 1\nWin %",     9),
    ("pre_team2_pct",  "Team 2\nWin %",     9),
    ("pre_winner",     "Predicted\nWinner", 22),
    ("pre_confidence", "Confidence",        10),
    ("pre_elo_diff",   "ELO\nDiff",          8),
    ("pre_form_diff",  "Form\nDiff",         8),
    ("pre_summary",    "LLM Summary",       40),
    ("pre_risk",       "Risk Flags",        30),
]

# Section: LIVE PREDICTIONS — overs 1–20 (20 × 2 cols = 40 cols)
# Each over: WIN% for batting team | score at that over
LIVE_START_COL = len(MATCH_COLS) + len(PRE_COLS) + 1  # col 17

# Section: ACTUAL RESULT (after over cols)
RESULT_COLS = [
    ("actual_winner",   "Actual\nWinner",   22),
    ("win_margin",      "Win Margin",       14),
    ("final_score_1",   "1st Inn\nScore",    9),
    ("final_score_2",   "2nd Inn\nScore",    9),
    ("mom",             "Player of\nMatch", 18),
]

# Section: ACCURACY
ACCURACY_COLS = [
    ("pre_correct",     "Pre-Match\nCorrect?",    10),
    ("live_o10_correct","Live O10\nCorrect?",     10),
    ("live_o15_correct","Live O15\nCorrect?",     10),
    ("pre_prob_correct","Pre %\nWas Right",        10),
    ("notes",           "Notes",                  30),
]

TOTAL_MATCH = len(MATCH_COLS)
TOTAL_PRE   = len(PRE_COLS)
TOTAL_LIVE  = 20 * 2   # 40 columns (win% + score per over)
TOTAL_RESULT = len(RESULT_COLS)
TOTAL_ACC    = len(ACCURACY_COLS)

RESULT_START_COL  = LIVE_START_COL + TOTAL_LIVE
ACCURACY_START_COL = RESULT_START_COL + TOTAL_RESULT

TOTAL_COLS = ACCURACY_START_COL + TOTAL_ACC - 1

# ── Build workbook ────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# Sheet 1: IPL Match Predictions
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Match Predictions"
ws.freeze_panes = "I3"   # freeze match info + row headers

# ── Row 1: Section headers (merged) ────────────────────────────────────────
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 36

sections = [
    (1, TOTAL_MATCH,                      "MATCH INFORMATION",      C_HEADER_MID),
    (TOTAL_MATCH+1, TOTAL_MATCH+TOTAL_PRE,"PRE-MATCH PREDICTION",   C_HEADER_GREEN),
    (LIVE_START_COL, LIVE_START_COL+TOTAL_LIVE-1, "LIVE WIN PROBABILITY (per over)", C_HEADER_ORANGE),
    (RESULT_START_COL, RESULT_START_COL+TOTAL_RESULT-1, "ACTUAL RESULT", C_HEADER_PURPLE),
    (ACCURACY_START_COL, ACCURACY_START_COL+TOTAL_ACC-1, "ACCURACY TRACKING", C_HEADER_TEAL),
]

for start, end, label, color in sections:
    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    cell = ws.cell(row=1, column=start)
    apply_cell(cell, label, bg=color, bold=True, font_color=C_WHITE, size=10)

# ── Row 2: Column headers ──────────────────────────────────────────────────
col = 1
for name, label, width in MATCH_COLS:
    apply_cell(ws.cell(2, col), label, bg=C_HEADER_MID, bold=True, font_color=C_WHITE)
    ws.column_dimensions[get_column_letter(col)].width = width
    col += 1

for name, label, width in PRE_COLS:
    apply_cell(ws.cell(2, col), label, bg=C_HEADER_GREEN, bold=True, font_color=C_WHITE)
    ws.column_dimensions[get_column_letter(col)].width = width
    col += 1

for over_num in range(1, 21):
    # Over sub-label row 1 (merged across 2 cols)
    # Win% column
    apply_cell(ws.cell(2, col), f"O{over_num}\nWin%", bg=C_HEADER_ORANGE, bold=True, font_color=C_WHITE, size=8)
    ws.column_dimensions[get_column_letter(col)].width = 6.5
    col += 1
    # Score column
    apply_cell(ws.cell(2, col), f"O{over_num}\nScore", bg=C_HEADER_ORANGE, bold=True, font_color=C_WHITE, size=8)
    ws.column_dimensions[get_column_letter(col)].width = 7
    col += 1

for name, label, width in RESULT_COLS:
    apply_cell(ws.cell(2, col), label, bg=C_HEADER_PURPLE, bold=True, font_color=C_WHITE)
    ws.column_dimensions[get_column_letter(col)].width = width
    col += 1

for name, label, width in ACCURACY_COLS:
    apply_cell(ws.cell(2, col), label, bg=C_HEADER_TEAL, bold=True, font_color=C_WHITE)
    ws.column_dimensions[get_column_letter(col)].width = width
    col += 1

# ── Pre-populate 60 match rows (IPL 2026 season) ──────────────────────────
for row_num in range(3, 63):
    bg = C_ROW_ALT if row_num % 2 == 0 else C_WHITE
    data_row = row_num - 2   # 1..60

    # Match info placeholders
    for c in range(1, TOTAL_MATCH + 1):
        apply_cell(ws.cell(row_num, c), None, bg=bg, center=True)

    # Pre-match prediction placeholders
    pre_start = TOTAL_MATCH + 1
    for c in range(pre_start, pre_start + TOTAL_PRE):
        apply_cell(ws.cell(row_num, c), None, bg=bg, center=True)

    # Win% cells: number format as percentage
    team1_pct_col = pre_start        # col 9
    team2_pct_col = pre_start + 1    # col 10
    ws.cell(row_num, team1_pct_col).number_format = "0%"
    ws.cell(row_num, team2_pct_col).number_format = "0%"

    # Live over columns
    for over_idx in range(20):
        win_col   = LIVE_START_COL + over_idx * 2
        score_col = LIVE_START_COL + over_idx * 2 + 1
        apply_cell(ws.cell(row_num, win_col),   None, bg=bg, center=True)
        apply_cell(ws.cell(row_num, score_col), None, bg=bg, center=True)
        ws.cell(row_num, win_col).number_format = "0%"

    # Result columns
    for c in range(RESULT_START_COL, RESULT_START_COL + TOTAL_RESULT):
        apply_cell(ws.cell(row_num, c), None, bg=bg, center=True)

    # Accuracy: formulas
    # Pre-match correct: 1 if predicted winner = actual winner
    pre_winner_col  = get_column_letter(TOTAL_MATCH + 3)   # col 11 = K
    actual_winner_col = get_column_letter(RESULT_START_COL) # first result col
    r = row_num

    pre_correct_col  = ACCURACY_START_COL
    o10_correct_col  = ACCURACY_START_COL + 1
    o15_correct_col  = ACCURACY_START_COL + 2
    prob_correct_col = ACCURACY_START_COL + 3
    notes_col        = ACCURACY_START_COL + 4

    pw_col   = get_column_letter(TOTAL_MATCH + 3)           # pre predicted winner
    aw_col   = get_column_letter(RESULT_START_COL)          # actual winner
    o10_col  = get_column_letter(LIVE_START_COL + 9*2 - 1)  # over 10 predicted winner (score col)
    o15_col  = get_column_letter(LIVE_START_COL + 14*2 - 1) # over 15

    # We'll store the live predicted winner as the score col value as a team name
    # and check against actual winner
    apply_cell(ws.cell(r, pre_correct_col),
               f'=IF(OR({aw_col}{r}="",{pw_col}{r}=""),"",IF({pw_col}{r}={aw_col}{r},"✓ Yes","✗ No"))',
               bg=bg, center=True)
    apply_cell(ws.cell(r, o10_correct_col), None, bg=bg, center=True)   # filled by API
    apply_cell(ws.cell(r, o15_correct_col), None, bg=bg, center=True)
    apply_cell(ws.cell(r, prob_correct_col), None, bg=bg, center=True)
    apply_cell(ws.cell(r, notes_col), None, bg=bg, center=False)

# ── Conditional formatting on win% cells ──────────────────────────────────
# Pre-match team1 win% — green when high
from openpyxl.formatting.rule import ColorScaleRule
pre_pct_col = get_column_letter(TOTAL_MATCH + 1)
ws.conditional_formatting.add(
    f"{pre_pct_col}3:{pre_pct_col}62",
    ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                   mid_type="num", mid_value=0.5, mid_color="FFEB84",
                   end_type="num", end_value=1, end_color="63BE7B")
)

# ════════════════════════════════════════════════════════════════════════════
# Sheet 2: Season Summary
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Season Summary")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 16
ws2.row_dimensions[1].height = 28

apply_cell(ws2.cell(1, 1), "IPL 2026 — Prediction Accuracy Summary",
           bg=C_HEADER_DARK, bold=True, font_color=C_WHITE, size=12, center=False)
ws2.merge_cells("A1:B1")

summary_rows = [
    ("Total Matches Played",          f"=COUNTA('Match Predictions'!A3:A62)"),
    ("Pre-Match Predictions Made",    f"=COUNTIF('Match Predictions'!{get_column_letter(TOTAL_MATCH+3)}3:{get_column_letter(TOTAL_MATCH+3)}62,\"<>\")"),
    ("Pre-Match Correct",             f"=COUNTIF('Match Predictions'!{get_column_letter(ACCURACY_START_COL)}3:{get_column_letter(ACCURACY_START_COL)}62,\"✓ Yes\")"),
    ("Pre-Match Accuracy %",          f"=IFERROR(B5/B4,\"-\")"),
    ("Live O10 Correct",              f"=COUNTIF('Match Predictions'!{get_column_letter(ACCURACY_START_COL+1)}3:{get_column_letter(ACCURACY_START_COL+1)}62,\"✓ Yes\")"),
    ("Live O10 Accuracy %",           f"=IFERROR(B6/B4,\"-\")"),
    ("Live O15 Correct",              f"=COUNTIF('Match Predictions'!{get_column_letter(ACCURACY_START_COL+2)}3:{get_column_letter(ACCURACY_START_COL+2)}62,\"✓ Yes\")"),
    ("Live O15 Accuracy %",           f"=IFERROR(B8/B4,\"-\")"),
]

for i, (label, formula) in enumerate(summary_rows, start=2):
    apply_cell(ws2.cell(i, 1), label, bg=C_LIGHT_BLUE if i%2==0 else C_WHITE, bold=False, center=False, size=10)
    cell = ws2.cell(i, 2)
    apply_cell(cell, formula, bg=C_LIGHT_BLUE if i%2==0 else C_WHITE, center=True, size=10)
    if "%" in label:
        cell.number_format = "0.0%"

# ── Save ──────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"✓ Created {OUTPUT_PATH}")
print(f"  Sheet 1: Match Predictions  ({TOTAL_COLS} columns × 62 rows)")
print(f"  Sheet 2: Season Summary")
print(f"\nColumn map:")
print(f"  A–{get_column_letter(TOTAL_MATCH)}   : Match information")
print(f"  {get_column_letter(TOTAL_MATCH+1)}–{get_column_letter(TOTAL_MATCH+TOTAL_PRE)} : Pre-match prediction")
print(f"  {get_column_letter(LIVE_START_COL)}–{get_column_letter(LIVE_START_COL+TOTAL_LIVE-1)} : Live win% per over")
print(f"  {get_column_letter(RESULT_START_COL)}–{get_column_letter(RESULT_START_COL+TOTAL_RESULT-1)} : Actual result")
print(f"  {get_column_letter(ACCURACY_START_COL)}–{get_column_letter(ACCURACY_START_COL+TOTAL_ACC-1)} : Accuracy tracking")
print(f"\nRun: python 05_create_excel_tracker.py")
print(f"Then start the API: uvicorn 04_api:app --host 0.0.0.0 --port 8000")
