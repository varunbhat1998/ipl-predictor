"""
04b_excel_writer.py
Adds /update-excel endpoint to the FastAPI app.
n8n calls this to write pre-match predictions, per-over live updates, and final results.

Import this in 04_api.py with:
    from excel_writer import router as excel_router
    app.include_router(excel_router)

Or run as standalone addition — see bottom of this file.
"""

from fastapi import APIRouter
from pydantic import BaseModel
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, threading

EXCEL_PATH = "ipl_predictions.xlsx"
LOCK = threading.Lock()   # prevent concurrent writes corrupting the file

router = APIRouter()

# ── Column constants (must match 05_create_excel_tracker.py) ──────────────
MATCH_COLS     = 8    # A–H
PRE_COLS       = 8    # I–P
LIVE_START_COL = MATCH_COLS + PRE_COLS + 1   # = 17 = col Q
RESULT_START   = LIVE_START_COL + 40         # after 20 overs × 2 cols = col BE (57)
ACCURACY_START = RESULT_START + 5            # col BJ (62)

# Column indices (1-based)
COL = {
    # Match info
    "match_id":       1,
    "season":         2,
    "date":           3,
    "team1":          4,
    "team2":          5,
    "venue":          6,
    "toss_winner":    7,
    "toss_decision":  8,
    # Pre-match
    "pre_team1_pct":  9,
    "pre_team2_pct":  10,
    "pre_winner":     11,
    "pre_confidence": 12,
    "pre_elo_diff":   13,
    "pre_form_diff":  14,
    "pre_summary":    15,
    "pre_risk":       16,
    # Result (RESULT_START = 57)
    "actual_winner":  RESULT_START,
    "win_margin":     RESULT_START + 1,
    "final_score_1":  RESULT_START + 2,
    "final_score_2":  RESULT_START + 3,
    "mom":            RESULT_START + 4,
    # Accuracy (ACCURACY_START = 62)
    "pre_correct":    ACCURACY_START,
    "live_o10":       ACCURACY_START + 1,
    "live_o15":       ACCURACY_START + 2,
    "prob_correct":   ACCURACY_START + 3,
    "notes":          ACCURACY_START + 4,
}

def over_win_pct_col(over_num: int) -> int:
    """Column index for win% at end of over N (1-based over)."""
    return LIVE_START_COL + (over_num - 1) * 2

def over_score_col(over_num: int) -> int:
    """Column index for score string at end of over N."""
    return LIVE_START_COL + (over_num - 1) * 2 + 1

# ── Helpers ────────────────────────────────────────────────────────────────
thin = openpyxl.styles.Side(style="thin", color="CCCCCC")
thin_border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

def write_cell(ws, row, col, value, bg=None, bold=False,
               color="000000", number_format=None, center=True):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = Font(name="Arial", size=9, bold=bold, color=color)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True
    )
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = thin_border
    if number_format:
        cell.number_format = number_format

def find_or_create_row(ws, match_id: str) -> int:
    """Find existing row for match_id or return next empty row."""
    for row in range(3, 200):
        val = ws.cell(row=row, column=COL["match_id"]).value
        if val is None or val == "":
            return row
        if str(val) == str(match_id):
            return row
    return 200

# ── Schemas ────────────────────────────────────────────────────────────────

class PreMatchWrite(BaseModel):
    match_id: str
    season: str
    date: str
    team1: str
    team2: str
    venue: str
    toss_winner: str
    toss_decision: str
    # prediction fields
    team1_win_probability: float
    team2_win_probability: float
    predicted_winner: str
    confidence: str
    elo_diff: Optional[float] = 0
    form_diff: Optional[float] = 0
    llm_summary: Optional[str] = ""
    risk_factors: Optional[list[str]] = []

class LiveOverWrite(BaseModel):
    match_id: str
    over_number: int           # 1–20
    batting_team_win_pct: float
    score_string: str          # e.g. "98/4"
    predicted_winner: str

class ResultWrite(BaseModel):
    match_id: str
    actual_winner: str
    win_margin: str            # e.g. "6 wickets" or "47 runs"
    first_innings_score: str   # e.g. "184/6"
    second_innings_score: str  # e.g. "138"
    player_of_match: Optional[str] = ""
    # accuracy fields — pass in from n8n after comparing
    live_o10_correct: Optional[str] = None   # "✓ Yes" / "✗ No"
    live_o15_correct: Optional[str] = None
    notes: Optional[str] = ""


# ── Endpoints ──────────────────────────────────────────────────────────────

@router.post("/update-excel/prematch")
def write_prematch(data: PreMatchWrite):
    """Write pre-match info + prediction to Excel. Creates new row."""
    with LOCK:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Match Predictions"]
        row = find_or_create_row(ws, data.match_id)
        bg = "F5F5F5" if row % 2 == 0 else "FFFFFF"
        bg_pre = "EBF5EB"

        # Match info
        write_cell(ws, row, COL["match_id"],      data.match_id,                bg=bg)
        write_cell(ws, row, COL["season"],         data.season,                  bg=bg)
        write_cell(ws, row, COL["date"],           data.date,                    bg=bg)
        write_cell(ws, row, COL["team1"],          data.team1,                   bg=bg)
        write_cell(ws, row, COL["team2"],          data.team2,                   bg=bg)
        write_cell(ws, row, COL["venue"],          data.venue,                   bg=bg, center=False)
        write_cell(ws, row, COL["toss_winner"],    data.toss_winner,             bg=bg)
        write_cell(ws, row, COL["toss_decision"],  data.toss_decision,           bg=bg)

        # Pre-match prediction
        write_cell(ws, row, COL["pre_team1_pct"],  data.team1_win_probability,   bg=bg_pre, number_format="0%")
        write_cell(ws, row, COL["pre_team2_pct"],  data.team2_win_probability,   bg=bg_pre, number_format="0%")
        write_cell(ws, row, COL["pre_winner"],     data.predicted_winner,        bg=bg_pre, bold=True, color="1D6B3A")
        write_cell(ws, row, COL["pre_confidence"], data.confidence,              bg=bg_pre)
        write_cell(ws, row, COL["pre_elo_diff"],   round(data.elo_diff or 0, 1), bg=bg_pre)
        write_cell(ws, row, COL["pre_form_diff"],  round(data.form_diff or 0, 3),bg=bg_pre)
        write_cell(ws, row, COL["pre_summary"],    data.llm_summary or "",       bg=bg_pre, center=False)
        risk_text = " | ".join(data.risk_factors or [])
        write_cell(ws, row, COL["pre_risk"],       risk_text,                    bg=bg_pre, center=False)

        wb.save(EXCEL_PATH)
    return {"status": "ok", "row": row, "match_id": data.match_id, "action": "prematch_written"}


@router.post("/update-excel/live")
def write_live_over(data: LiveOverWrite):
    """Write win probability for one over. Called by n8n after each over."""
    if not 1 <= data.over_number <= 20:
        return {"status": "error", "message": "over_number must be 1–20"}
    with LOCK:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Match Predictions"]
        row = find_or_create_row(ws, data.match_id)

        win_col   = over_win_pct_col(data.over_number)
        score_col = over_score_col(data.over_number)

        # Win% — colour code: green if ≥60%, red if ≤40%, yellow otherwise
        pct = data.batting_team_win_pct / 100 if data.batting_team_win_pct > 1 else data.batting_team_win_pct
        if pct >= 0.60:
            live_bg = "C6EFCE"   # green
        elif pct <= 0.40:
            live_bg = "FFC7CE"   # red
        else:
            live_bg = "FFEB9C"   # yellow

        write_cell(ws, row, win_col,   pct,                  bg=live_bg, number_format="0%")
        write_cell(ws, row, score_col, data.score_string,    bg="FCE4D6")

        # Over 10 and 15: also write predicted winner to accuracy cols
        if data.over_number == 10:
            write_cell(ws, row, COL["live_o10"], None, bg="FFFFFF")   # placeholder (formula fills)
        if data.over_number == 15:
            write_cell(ws, row, COL["live_o15"], None, bg="FFFFFF")

        wb.save(EXCEL_PATH)
    return {"status": "ok", "row": row, "over": data.over_number, "win_pct": round(pct, 3)}


@router.post("/update-excel/result")
def write_result(data: ResultWrite):
    """Write actual match result + accuracy checks after match ends."""
    with LOCK:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Match Predictions"]
        row = find_or_create_row(ws, data.match_id)
        bg = "EAD1DC"   # light purple for result section

        write_cell(ws, row, COL["actual_winner"],  data.actual_winner,          bg=bg, bold=True, color="4B0082")
        write_cell(ws, row, COL["win_margin"],     data.win_margin,             bg=bg)
        write_cell(ws, row, COL["final_score_1"],  data.first_innings_score,    bg=bg)
        write_cell(ws, row, COL["final_score_2"],  data.second_innings_score,   bg=bg)
        write_cell(ws, row, COL["mom"],            data.player_of_match or "",  bg=bg)

        # Accuracy columns
        if data.live_o10_correct:
            c = "C6EFCE" if "Yes" in data.live_o10_correct else "FFC7CE"
            write_cell(ws, row, COL["live_o10"], data.live_o10_correct, bg=c, bold=True)
        if data.live_o15_correct:
            c = "C6EFCE" if "Yes" in data.live_o15_correct else "FFC7CE"
            write_cell(ws, row, COL["live_o15"], data.live_o15_correct, bg=c, bold=True)
        if data.notes:
            write_cell(ws, row, COL["notes"], data.notes, bg="FFFFFF", center=False)

        # Highlight the pre-match correct cell based on formula result
        # (formula is already in the cell from creation — just ensure it's there)

        wb.save(EXCEL_PATH)
    return {"status": "ok", "row": row, "match_id": data.match_id, "actual_winner": data.actual_winner}


@router.get("/excel/status")
def excel_status():
    """Returns how many matches have been written."""
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["Match Predictions"]
    count = 0
    for row in range(3, 200):
        if ws.cell(row=row, column=1).value:
            count += 1
        else:
            break
    wb.close()
    return {"matches_written": count, "excel_path": EXCEL_PATH}
